import functions_framework
import io
from PIL import Image
import looker_sdk
import os
import pandas as pd
import numpy as np
from datetime import datetime
import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from google.cloud import storage
from google.cloud import bigquery
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import requests
import base64
import sib_api_v3_sdk
from sib_api_v3_sdk.rest import ApiException
from google.cloud import secretmanager_v1
import jinja2
import datetime as dt
from variables import *


start_time = time.time()
def suffix(d):
    return 'th' if 11<=d<=13 else {1:'st',2:'nd',3:'rd'}.get(d%10, 'th')

def custom_strftime(format, t):
    return t.strftime(format).replace('{S}', str(t.day) + suffix(t.day))

def dollar_sign(df,float_columns):
    
    # Add $ sign to float columns
    for col in float_columns:
        df[col] = df[col].apply(lambda x: f'${x:,.2f}')
    return df
    
def percentage_sign(df,float_columns):
    
    # Add $ sign to float columns
    for col in float_columns:
        df[col] = df[col].apply(lambda x: f'{x:,.2f}%')
    return df

def convert_date_columns(df, date_columns):
    
    for col in date_columns:
        if col in df.columns:
            try:
                df[col] = pd.to_datetime(df[col], format='%Y-%m-%d').dt.strftime('%m-%d-%Y')
            except ValueError:
                # If conversion fails, print a message and skip the column
                print(f"Column '{col}' could not be converted to date format.")
    return df

def check_log_file_in_gcs(bucket_name):
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    
    blobs = bucket.list_blobs()
    
    # Check for any .log files
    for blob in blobs:
        if blob.name.endswith('.log'):
            return True  # At least one .log file exists
    
    return False  # No .log files found

def query_bigquery():
    # Initialize BigQuery client
    client = bigquery.Client()
    result = False
    query = """
        SELECT * FROM reporting.dw_pipeline_log WHERE run_finished_time >= DATETIME(CONCAT(CAST(CURRENT_DATE("Asia/Calcutta") AS STRING), ' 12:30:00')) LIMIT 1
    """
    
    # Run the query
    query_job = client.query(query)
    
    # Get the results
    results = query_job.result()  # Waits for the job to complete
    for row in results:
        date = row['run_finished_time'].date()
        print(date)
        if date == datetime.now().date():
            result = True
    return result

def get_secret(secret_id):

    client = secretmanager_v1.SecretManagerServiceClient()
    name = f"projects/{secret_project_id}/secrets/{secret_id}/versions/latest"
    response = client.access_secret_version(name=name)
    secret_data = response.payload.data.decode("UTF-8")
    try:
        secret_dict = json.loads(secret_data)
        return secret_dict
    except json.JSONDecodeError:
        # If it's not JSON, return the data as a plain string
        return secret_data
    
date_for_mail = custom_strftime('%B {S}, %Y', datetime.now())
file_path='/tmp/report.xlsx'

file_name = f'Pipeline Dashboard - {date_for_mail}'

storage_client = storage.Client(project = project_id)

bucket = storage_client.get_bucket(destination_bucket)
get_bucket = storage_client.get_bucket(destination_bucket)

values_list = []


def lambda_handler(response):
    log_file_exists = check_log_file_in_gcs(log_bucket_name)
    pipeline_ran_today = query_bigquery()
    if log_file_exists or not pipeline_ran_today:
        print('Either files are missing or Pipeline did not ran. Cannot send emails')
        return {
        'statusCode': 500,
        'body': json.dumps('Either files are missing or Pipeline did not ran. Cannot send emails')
    }
    else: # log file doesn't exist and pipeline ran today
        print("No .log files found in the bucket.")
    open(file_path, "w").close()
    print(f'{file_path} erased')
    open('/tmp/view.png', "w").close()
    print(f'/tmp/view.png erased')
    open('/tmp/view(1).png', "w").close()
    print(f'/tmp/view(1).png erased')
    pipeline_dashboard()
    open(file_path, "w").close()
    print(f'{file_path} erased')
    open('/tmp/view.png', "w").close()
    print(f'/tmp/view.png erased')
    open('/tmp/view(1).png', "w").close()
    print(f'/tmp/view(1).png erased')
    print("--- %s seconds ---" % (time.time() - start_time))

    return {
        'statusCode': 200,
        'body': json.dumps('pipeline report sent successfully')
    }

def pipeline_dashboard():

    looker_creds = get_secret(secret_name['looker_creds'])
    os.environ['LOOKERSDK_BASE_URL'] = looker_creds['LOOKERSDK_BASE_URL']
    os.environ['LOOKERSDK_CLIENT_ID'] = looker_creds['LOOKERSDK_CLIENT_ID']
    os.environ['LOOKERSDK_CLIENT_SECRET'] = looker_creds['LOOKERSDK_CLIENT_SECRET']
    
    sdk = looker_sdk.init40()

    #all sheet
    response = sdk.run_look(str(look_ids['all sheet']), "csv")
    result = pd.read_csv(io.StringIO(response), header = None)
    result = result.where(pd.notna(result), None)
    result.replace({'Pre Post Loans Product Type': 'Product Type', 'Pre Post Loans Originator': 'Originator'}, inplace=True)
    result.loc[0] = result.loc[0].fillna('Total')

    unique_columns = tuple(list(result.iloc[0].unique())[1:])

    new_header = result.iloc[0] 
    result = result[1:]  
    result.columns = new_header  

    # Reset index if needed
    result.reset_index(drop=True, inplace=True)

    #POST Close DoD Difference
    response = sdk.run_look(str(look_ids['POST Close DoD Difference']), "csv")
    post_dod_df = pd.read_csv(io.StringIO(response))
    post_dod_df['Post Loans Originator'] = post_dod_df['Post Loans Originator'].fillna('Grand Total')
    post_dod_df = post_dod_df.where(pd.notna(post_dod_df), None)
    post_dod_df.columns = [col if 'Post Loans' not in col else col.replace('Post Loans ','') for col in post_dod_df.columns]


    # PRE Close DoD Difference
    response = sdk.run_look(str(look_ids['PRE Close DoD Difference']), "csv")
    pre_dod_df = pd.read_csv(io.StringIO(response))
    pre_dod_df['Pre Loans Originator'] = pre_dod_df['Pre Loans Originator'].fillna('Grand Total')
    pre_dod_df = pre_dod_df.where(pd.notna(pre_dod_df), None)
    pre_dod_df.columns = [col if 'Pre Loans' not in col else col.replace('Pre Loans ','') for col in pre_dod_df.columns]



    # PRE POST Close DoD Difference
    response = sdk.run_look(str(look_ids['PRE POST Close DoD Difference']), "csv")
    pre_post_close_loans = pd.read_csv(io.StringIO(response))
    pre_post_close_loans['Pre Post Loans Originator'] = pre_post_close_loans['Pre Post Loans Originator'].fillna('Grand Total')
    pre_post_close_loans = pre_post_close_loans.where(pd.notna(pre_post_close_loans), None)
    pre_post_close_loans.columns = [col if 'Pre Post Loans' not in col else col.replace('Pre Post Loans ','') for col in pre_post_close_loans.columns]


    # PRE Close Submission
    response = sdk.run_look(str(look_ids['PRE Close Submission']), "csv")
    pre_close = pd.read_csv(io.StringIO(response))
    pre_close = pre_close.where(pd.notna(pre_close), None)
    pre_close.columns = [col if 'Pre Close Submission' not in col else col.replace('Pre Close Submission Loans ','') for col in pre_close.columns]



    # Pre Channel % UPB
    pre_channel = pre_close.groupby('PRE Chanel')['% UPB'].sum().mul(100).round(2).reset_index()


    # POST Close Submission
    response = sdk.run_look(str(look_ids['POST Close Submission']), "csv")
    post_close = pd.read_csv(io.StringIO(response))
    post_close = post_close.where(pd.notna(post_close), None)
    post_close.columns = [col if 'Post Close Submission' not in col else col.replace('Post Close Submission Loans ','') for col in post_close.columns]

    # Post Channel % UPB
    post_channel = post_close.groupby('POST Channel')['% UPB'].sum().mul(100).round(2).reset_index()



    # agg channel % upb
    agg_pre = pre_close[['Current Principal Balance','PRE Chanel']]
    agg_pre.rename(columns={'Current Principal Balance': 'cpb', 'PRE Chanel': 'Aggregate Channel'}, inplace=True)
    agg_post = post_close[['Current Principal Balance','POST Channel']]
    agg_post.rename(columns={'Current Principal Balance': 'cpb', 'POST Channel': 'Aggregate Channel'}, inplace=True)
    agg = pd.concat([agg_pre,agg_post])
    agg = agg.where(pd.notna(agg), None)
    agg_channel = agg.groupby('Aggregate Channel')['cpb'].sum().reset_index()
    agg_channel['% UPB'] = (agg_channel['cpb']/sum(agg_channel['cpb'])*100).round(2)
    agg_channel = agg_channel[['Aggregate Channel','% UPB']]

    #agg %upb format
    agg_channel = percentage_sign(agg_channel, ['% UPB'])
    agg_channel_html = agg_channel.to_html(index=False)

    #post dod format    
    post_dod_df = dollar_sign(post_dod_df,['POST Close Previous Day','POST Close','POST Delta'])
    values_list = post_dod_df[post_dod_df['Originator'] == 'Grand Total'].values.tolist()[0][1:]


    #pre dod format    
    pre_dod_df = dollar_sign(pre_dod_df,['PRE Close Previous Day','PRE Close','PRE Delta'])
    values_list.extend(pre_dod_df[pre_dod_df['Originator'] == 'Grand Total'].values.tolist()[0][1:])

    #pre post dod format
    pre_post_close_loans = dollar_sign(pre_post_close_loans,['POST & PRE Close Previous Day','POST & PRE Close','Delta'])
    values_list.extend(pre_post_close_loans[pre_post_close_loans['Originator'] == 'Grand Total'].values.tolist()[0][1:])

    #pre submission format
    pre_close = dollar_sign(pre_close, ['Initial Loan Amount','Maximum Loan Amount','Current Principal Balance','Total Principal Balance'])
    pre_close = percentage_sign(pre_close, ['% UPB'])
    pre_df_html = pre_close.to_html(index=False, na_rep='')

    #pre channel format
    pre_channel = percentage_sign(pre_channel, ['% UPB'])
    pre_channel_html = pre_channel.to_html(index=False)


    #post submission format
    post_close = dollar_sign(post_close, ['Initial Loan Amount','Maximum Loan Amount','Current Principal Balance','Total Principal Balance'])
    post_close = percentage_sign(post_close, ['% UPB'])
    post_df_html = post_close.to_html(index=False)

    #post channel format
    post_channel = percentage_sign(post_channel, ['% UPB'])
    post_channel_html = post_channel.to_html(index=False)



    #Loan Info sheet
    response = sdk.run_look(str(look_ids['Loan Info']), "csv", limit=7000)
    loan_info = pd.read_csv(io.StringIO(response))
    loan_info = loan_info.where(pd.notna(loan_info), None)
    loan_info.columns = [col if 'Loan Information' not in col else col.replace('Loan Information ','') for col in loan_info.columns]
    loan_info = loan_info[['Originator', 'Toorak Loan Id', 'Fund Type', 'Regulartoorakyield',
    'prepay Penalty Type', 'Toorak Rate', 'Interest Rate',
    'Originator Loan Id', 'Toorak Product', 'Loan Process Type',
    'Loan State', 'Property Type', 'Loan Stage', 'Due Diligence Grade',
    'Created Date', 'Rate Lock Expiry Date', 'Pricing Date',
    'Pre Submit Date', 'Current Submit Date', 'Convert Pre Date',
    'Convert Post Date', 'Initial Review', 'Due Diligence Review Date',
    'Final Review Date', 'Purchase Initiated Date', 'Target Purchase Date',
    'Purchased Date', 'Closing Date', '% Occupied', 'Appraisal Noi',
    'Loan Purpose', 'Number of Units', 'Toorak Ncf', 'Financing Type',
    'Final Toorak Price', 'Property Address', 'Current Principal Balance',
    'Original Maximum Loan Amount', 'Toorak Lending Amount',
    'Assigned Partner']]
    loan_info = convert_date_columns(loan_info , ['Created Date', 'Rate Lock Expiry Date', 'Pricing Date',
        'Pre Submit Date', 'Current Submit Date', 'Convert Pre Date',
        'Convert Post Date', 'Initial Review', 'Due Diligence Review Date',
        'Final Review Date', 'Purchase Initiated Date', 'Target Purchase Date',
        'Purchased Date', 'Closing Date'])
    loan_info = dollar_sign(loan_info,['Current Principal Balance','Original Maximum Loan Amount', 'Toorak Lending Amount'])
    

    with pd.ExcelWriter(file_path) as writer:

        result.to_excel(writer, sheet_name = 'Pipeline Dashboard Export')
        post_dod_df.to_excel(writer, sheet_name = 'POST Close DoD Difference', index=False)
        pre_dod_df.to_excel(writer, sheet_name = 'PRE Close DoD Difference', index=False)
        pre_post_close_loans.to_excel(writer, sheet_name = 'PRE POST Close DoD Difference', index=False)
        pre_close.to_excel(writer, sheet_name = 'PRE Close Submission', index=False)
        post_close.to_excel(writer, sheet_name = 'POST Close Submission', index=False)
        pre_channel.to_excel(writer, sheet_name = 'Pre Channel % UPB', index=False)
        post_channel.to_excel(writer, sheet_name = 'Post Channel % UPB', index=False)
        agg_channel.to_excel(writer, sheet_name = 'Aggregate Channel % UPB', index=False)
        loan_info.to_excel(writer, sheet_name = 'Loan Info Export')
        
    
    #loans by originator 
    
    response = sdk.run_look(str(look_ids['loans by originator']), "png")
    image = Image.open(io.BytesIO(response))
    image.save('/tmp/view.png')

    ws_name = file_path
    wb = openpyxl.load_workbook(file_path)
    wb.create_sheet("Data Visualisation",0)
    ws = wb["Data Visualisation"]
    img = openpyxl.drawing.image.Image('/tmp/view.png')
    img.anchor = 'B2'
    img.width = 850
    ws.add_image(img)
    
    #loans priciple balance
    
    response = sdk.run_look(str(look_ids['loans priciple balance']), "png")
    image = Image.open(io.BytesIO(response))
    image.save('/tmp/view(1).png')

    ws = wb["Data Visualisation"]
    img = openpyxl.drawing.image.Image('/tmp/view(1).png')
    img.anchor = 'Q2'
    img.width = 850
    img.height = 900
    ws.add_image(img)
    
    wb.save(ws_name)

    wb = openpyxl.load_workbook(file_path)
    ws = wb['Pipeline Dashboard Export']

    columns_to_merge_list = dict()
    # Define the header you want to merge into
    for header_to_merge in unique_columns:
        # header_to_merge = 'Bridge(1-4 Units)'
        
        # Get the column headers from the first row
        headers = [cell.value for cell in ws[1]]
        headers[0] = 'empty'
        
        # # Identify columns that need to be merged
        columns_to_merge_list[header_to_merge] = [i + 1 for i, header in enumerate(headers) if header.startswith(header_to_merge)]
        # Merge cells if there are multiple columns with the same header prefix
    if columns_to_merge_list:
        for columns_to_merge in columns_to_merge_list:
            start_col = columns_to_merge_list[columns_to_merge][0]
            end_col = columns_to_merge_list[columns_to_merge][-1]
        
            # Merge cells in the first row
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        
            # Set the merged cell value and alignment
            merged_cell = ws.cell(row=1, column=start_col)
            merged_cell.value = columns_to_merge
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')



    # Save the modified Excel file
    wb.save(file_path)

    with open(file_path, 'rb') as f:
        txt=f.read()
        # print('read the file')
        blob = bucket.blob(f'looker_report_emails/Pipeline_Dashboard_Email.xlsx')
        blob.upload_from_string(txt)
        attachments_ = txt

    sender_ = sender_email
    recipients_ = email_recipients
    print(recipients_)
    title_ = file_name
    text_ = 'this is test'
    body_ = """<html>
              <head>
                <ul id="ul-img" style="list-style-type: none;margin: 0;padding: 0;background-color: #0fcbef;width: 100%;height: 60px;border-radius: 4px;background-image: -webkit-linear-gradient(97deg, #0fcbef 4%, #1071ee 90%) !important;">
                  <div>
                    <img src="https://file-management-dev-tl-ue1-s3.s3.amazonaws.com/tooraklogo.png" />
                  </div>
                </ul>
            </head>
            <style> 
             table, th, td {{ border: 1px solid black; border-collapse: collapse; }}
              th, td {{ padding: 5px; }}
            </style>
            <body>
            <br>Hi All,<br><br>Please find attached Pipeline Dashboard as on date_for_sending_email.<br>Please let us know if you face any issues.<br>
            <br><br>
            <br>PRE CLOSE SUBMISSIONS<br>
            <br>pre_close_data<br>
            <br><br>
            <br>PRE CLOSE DELTA<br>
            <br><br>
            <table 
                <tr>
                    <th>Originator</th>
                    <th>PRE CLOSE Previous day</th>
                    <th>PRE CLOSE Current day</th>
                    <th>PRE CLOSE Delta</th>
                </tr>
                <tr>
                    <td>{}</td>
                    <td>{}</td>
                    <td>{}</td>
                    <td>{}</td>
                </tr>
            </table>
            <br><br>
            <br>Pre Channel % UPB<br>
            <br>pre_channel_html<br>
            <br><br>
            <br>POST CLOSE SUBMISSIONS<br>
            <br>post_close_data<br>
            <br>POST CLOSE Delta<br>
            <br><br>
            <table 
                <tr>
                    <th>Originator</th>
                    <th>POST CLOSE Previous day</th>
                    <th>POST CLOSE Current day</th>
                    <th>POST CLOSE Delta</th>
                </tr>
                <tr>
                    <td>{}</td>
                    <td>{}</td>
                    <td>{}</td>
                    <td>{}</td>
                </tr>
            </table>
            <br><br>
            <br>Post Channel % UPB<br>
            <br>post_channel_html<br>
            <br><br>
            <br>PRE/POST CLOSE DELTA<br>
            <br><br>
            <table 
                <tr>
                    <th>Originator</th>
                    <th>Aggregate Pre/Post CLOSE Previous day</th>
                    <th>Aggregate Pre/Post CLOSE Current day</th>
                    <th>Delta</th>
                </tr>
                <tr>
                    <td>{}</td>
                    <td>{}</td>
                    <td>{}</td>
                    <td>{}</td>
                </tr>
            </table>
            <br><br>
            <br>Aggregate Channel % UPB<br>
            <br>agg_channel_html<br>
            <br><br>
            <br>Thanks,<br>Toorak.<br>
            <img src='https://static.wixstatic.com/media/74c4f9_20406c39aa61491d83b0ccec086c6feb~mv2_d_4500_4500_s_4_2.png/v1/fit/w_1000%2Ch_1000%2Cal_c/file.png' width="250" height="190">
            </body>
            </html>""".format('Grand Total',values_list[3],values_list[4],values_list[5],'Grand Total',values_list[0],values_list[1],values_list[2],'Grand Total',values_list[6],values_list[7],values_list[8])
   
    body_ = body_.replace('date_for_sending_email', date_for_mail).replace('post_close_data',post_df_html).replace('pre_close_data',pre_df_html).replace('post_channel_html',post_channel_html).replace('pre_channel_html',pre_channel_html).replace('agg_channel_html',agg_channel_html)
    response=send_mail(sender_, recipients_, title_, text_, body_, attachments_,file_name)
    



def send_mail(sender, recipients, title, text, html, attachments,filename):
    """
    Send email to recipients. Sends one mail to all recipients.
    The sender needs to be a verified email in SES.
    """
    #msg = create_multipart_message(sender, recipients, title, text, html, attachments)
    
#     ses_client = boto3.client('ses',
#                              )

    #ses_client = boto3.client('ses',region_name='us-east-1')
    configuration = sib_api_v3_sdk.Configuration()
    configuration.api_key['api-key'] = get_secret(secret_name['email_api'])
    api_instance = sib_api_v3_sdk.TransactionalEmailsApi(sib_api_v3_sdk.ApiClient(configuration)) 
    #html_title="<html><body><h1>{{title}}</h1></body></html>" 
    headers = {"Content-Disposition":"Attachments"}

    encoded_string = base64.b64encode(attachments)
    base64_message = encoded_string.decode('utf-8')
    filename+=".xlsx"
    attachment = [{"content":base64_message,"name":filename}]
    send_smtp_email = sib_api_v3_sdk.SendSmtpEmail(to=recipients, html_content=html, sender=sender, subject=title,headers=headers,attachment=attachment)

    #return ses_client.send_raw_email(Source=sender,Destinations=recipients,RawMessage={'Data': msg.as_string()})
    try:
        api_response = api_instance.send_transac_email(send_smtp_email)
        return api_response
    except ApiException as e:
        print("Exception when calling SMTPApi->send_transac_email: %s\n" % e)
        return e
    print('mail sent')
