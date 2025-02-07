from variables import * 
from send_email import *
import io
from PIL import Image
import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from google.cloud import storage
import json
from openpyxl.styles import Font
import jinja2
from uk_emailer.variables import *

def modify_uk_excel_cells(ws, rows_to_modify):
    """
    Modify specified rows in the Excel worksheet.

    Parameters:
    ws (Worksheet): The worksheet to modify.
    rows_to_modify (list): List of row numbers to modify.
    """
    for row in rows_to_modify:
        for col in range(2, 54):  # B is column 2 and BA is column 53
            cell = ws.cell(row=row, column=col)
            
            if isinstance(cell.value, str):  # Ensure the cell contains a string
                # if row in [4, 5, 6, 7, 13, 14]:   removing Pound sign on 6th Feb 2024
                #     cell.value = f"£{cell.value}"
                    
                if row in [47, 48, 49, 50]:
                    try:
                        numeric_value = float(cell.value)  # Convert to float
                        formatted_value = int(numeric_value * 100) # Multiply by 100 and round to 2 decimal places
                        cell.value = f"{formatted_value}%"  # Append '%'
                    except ValueError:
                        pass  # If conversion fails, keep original value
                
                # For row 24: Append percentage sign
                if row in [31,41]:
                    cell.value = f"{cell.value}%"

        pound_format = '#,##0.00'

        specific_rows = [4, 5, 6, 7, 13, 14, 21, 23, 24, 37, 38, 39, 40]

# Apply Pound format and convert values to float to specific rows and columns (B to BA)
    for row in ws.iter_rows(min_row=5, max_row=50, min_col=2, max_col=53):
        for cell in row:
            if cell.row in specific_rows:  # Check if the current row is in the list
                if isinstance(cell.value, str):  # If the value is a string
                    # Remove any non-numeric characters (like currency symbols and commas)
                    cell.value = ''.join(e for e in cell.value if e.isdigit() or e == '.')
            
                try:
                    # Attempt to convert the value to float (this will work for strings that are numeric)
                    cell.value = float(cell.value)
                except ValueError:
                    # If conversion fails (non-numeric value), leave the value as is
                    pass
                
                # Apply the pound format after conversion
                cell.number_format = pound_format
                
            if cell.row in [22, 25]:
                if isinstance(cell.value, str):  # If the value is a string
                    try:
                        # Convert the string to an integer if possible
                        cell.value = int(cell.value)
                    except ValueError:
                        # If conversion fails, leave the value as is
                        pass

def auto_adjust_column_width(ws):
    """
    Adjusts the width of all columns, setting a fixed width for the first column
    and adjusting the rest based on the data they contain.
    """
    # Set a fixed width for Column A (adjust the value as needed)
    ws.column_dimensions['A'].width = 30  # You can change 30 to any fixed width you prefer

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        # Skip Column A since its width is already fixed
        if col_letter == 'A':
            continue

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        ws.column_dimensions[col_letter].width = max_length + 2

def uk_weekly_report(file_name, sdk, email_api, bucket, get_bucket):

    response = sdk.run_look(str(uk_look_ids['table_1']), "csv")
    df = pd.read_csv(io.StringIO(response))
    df = df.drop(df.index[1])
    df = df.drop(df.columns[2], axis=1)
    Purchases = df.drop(df.columns[1], axis=1)

    response = sdk.run_look(str(uk_look_ids['table_2']), "csv")
    df = pd.read_csv(io.StringIO(response))
    Draws_and_Payoffs = df.drop(df.index[1])


    response = sdk.run_look(str(uk_look_ids['table_3']), "csv")
    df = pd.read_csv(io.StringIO(response))
    df = df.drop(df.index[1])
    Totals = df.drop(df.columns[1], axis=1)


    response = sdk.run_look(str(uk_look_ids['table_4']), "csv")
    df = pd.read_csv(io.StringIO(response))
    df = df.drop(df.index[1])
    Net_WAC = df.loc[:, ~df.columns.str.endswith('.1')]


    response = sdk.run_look(str(uk_look_ids['table_5']), "csv")
    df = pd.read_csv(io.StringIO(response))
    df = df.drop(df.index[1])
    df = df.drop(df.columns[0], axis=1)
    Four_Week_Rolling = df.loc[:, ~df.columns.str.endswith('.1')]
    

    response = sdk.run_look(str(uk_look_ids['table_6']), "csv")
    df = pd.read_csv(io.StringIO(response))
    df = df.drop(df.index[1])
    Rolling_4_4_Week_Delta = df.drop(df.columns[1], axis=1)
    
    with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
        Purchases.to_excel(writer, startrow=1, startcol=0, index = False)
        Draws_and_Payoffs.to_excel(writer, startrow=10, startcol=0, index = False)
        Totals.to_excel(writer, startrow=18, startcol=0, index = False)
        Net_WAC.to_excel(writer, startrow=28, startcol=0, index = False)
        Four_Week_Rolling.to_excel(writer, startrow=34, startcol=0, index = False)
        Rolling_4_4_Week_Delta.to_excel(writer, startrow=44, startcol=0, index = False)


    response = sdk.run_look(str(uk_look_ids['graph_1']), "png", image_width=1500, image_height=500)
    image = Image.open(io.BytesIO(response))
    image.save('view.png')
    ws_name = 'output.xlsx'
    file_path = 'output.xlsx'
    wb = openpyxl.load_workbook(file_path)
    wb.create_sheet("Data Visualisation",1)
    ws = wb["Data Visualisation"]
    img = openpyxl.drawing.image.Image('view.png')
    img.anchor = 'B4'
    img.width = 1200
    ws.add_image(img)

    response = sdk.run_look(str(uk_look_ids['graph_2']), "png", image_width=1500, image_height=500)
    image = Image.open(io.BytesIO(response))
    image.save('view_1.png')
    img = openpyxl.drawing.image.Image('view_1.png')
    img.anchor = 'B35'
    img.width = 1200
    ws.add_image(img)


    response = sdk.run_look(str(uk_look_ids['graph_3']), "png", image_width=1500, image_height=500)
    image = Image.open(io.BytesIO(response))
    image.save('view_2.png')
    img = openpyxl.drawing.image.Image('view_2.png')
    img.anchor = 'B66'
    img.width = 1200
    ws.add_image(img)

    response = sdk.run_look(str(uk_look_ids['graph_4']), "png", image_width=1500, image_height=500)
    image = Image.open(io.BytesIO(response))
    image.save('view_3.png')
    img = openpyxl.drawing.image.Image('view_3.png')
    img.anchor = 'B97'
    img.width = 1200
    ws.add_image(img)

    response = sdk.run_look(str(uk_look_ids['graph_5']), "png", image_width=1500, image_height=500)
    image = Image.open(io.BytesIO(response))
    image.save('view_4.png')
    img = openpyxl.drawing.image.Image('view_4.png')
    img.anchor = 'B128'
    img.width = 1200
    ws.add_image(img)

    ws['F33'].font = Font(size=18, bold=True) 
    ws['F33'] = "Bridge Volume" 
    ws['F2'].font = Font(size=18, bold=True) 
    ws['F2'] = "Purchase Volume" 
    ws['F64'].font = Font(size=18, bold=True) 
    ws['F64'] = "Draws Volume" 
    ws['F95'].font = Font(size=18, bold=True) 
    ws['F95'] = "Payoffs Volume" 
    ws['F126'].font = Font(size=18, bold=True) 
    ws = wb["Sheet1"]
    ws.freeze_panes = "B3"
    ws['F126'] = "NetWAC (Toorak Yield)" 
    ws['A2'] = "Week Number"
    ws['A3'] = "Purchases"
    ws['A11'] = "Week Number"
    ws['A12'] = "Draws & Payoffs"
    ws['A19'] = "Week Number"
    ws['A20'] = "Totals"
    ws['A29'] = "Week Number"
    ws['A30'] = "Net WAC (Toorak Yield)"
    ws['A35'] = "Week Number"
    ws['A36'] = "4 Week Rolling"
    ws['A45'] = "Week Number"
    ws['A46'] = "Rolling 4 4 Week Delta"

    rows_to_modify = [31, 41, 47, 48, 49, 50] # Since Pound sign has been removed no need to add row 4,5,6,7,13,14 to this list
    #  [4, 5, 6, 7, 13, 14]
    modify_uk_excel_cells(ws, rows_to_modify)
    # Auto adjust column widths after modifying the cells
    auto_adjust_column_width(ws)

    wb.save(ws_name)

    with open(file_path, 'rb') as f:
        txt=f.read()
        # print('read the file')
        blob = bucket.blob(f'looker_report_emails/{file_name}.xlsx')
        blob.upload_from_string(txt)
        attachments_ = txt
        
    text_ = 'this is test'
    body_ = """<html>
                <head>
                <ul id="ul-img" style="list-style-type: none;margin: 0;padding: 0;background-color: #0fcbef;width: 100%;height: 60px;border-radius: 4px;background-image: -webkit-linear-gradient(97deg, #0fcbef 4%, #1071ee 90%) !important;">
                    <div>
                    <img src="https://file-management-dev-tl-ue1-s3.s3.amazonaws.com/tooraklogo.png" />
                    </div>
                </ul>
            </head>
            <body>
            <br>Hi All,<br><br>Please find attached report name.<br>Please let us know if you face any issues.<br><br>Thanks,<br>Toorak.<br>
            <img src='https://static.wixstatic.com/media/74c4f9_20406c39aa61491d83b0ccec086c6feb~mv2_d_4500_4500_s_4_2.png/v1/fit/w_1000%2Ch_1000%2Cal_c/file.png' width="250" height="190">
            </body>
            </html>"""
    body_ = body_.replace('report name',file_name)
        
    response = send_mail(sender_email, uk_email_recipients, uk_email_cc, text_, body_, file_name, attachments_, file_name, email_api)


