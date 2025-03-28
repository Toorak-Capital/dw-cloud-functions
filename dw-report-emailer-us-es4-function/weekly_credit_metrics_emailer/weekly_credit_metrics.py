from variables import * 
import io
from send_email import *
from PIL import Image
import looker_sdk
import os
import pandas as pd
import numpy as np
from datetime import *
import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import json
import requests
import base64
import jinja2
import datetime as dt
import re
from openpyxl.styles import Border, Side
from openpyxl.utils import column_index_from_string
from weekly_credit_metrics_emailer.variables import *

def convert_date_columns(df, date_columns):
    
    for col in date_columns:
        if col in df.columns:
            try:
                df[col] = pd.to_datetime(df[col], format='%Y-%m-%d').dt.strftime('%m-%d-%Y')
            except ValueError:
                # If conversion fails, print a message and skip the column
                print(f"Column '{col}' could not be converted to date format.")
    return df

def modify_excel_cells(ws, rows_to_modify):
    for row in rows_to_modify:
        for col in range(2, 54):  # B is column 2 and BA is column 53
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):  # Ensure the cell contains a string
                # Remove the '.x' part if it exists
                new_value = cell.value.split('.')[0]  # Keep the part before '.'
                cell.value = new_value


def multiply_and_format(ws, row_numbers, num, start_col=2, end_col=53):
    """
    Multiplies each cell's value in the specified rows by 100,
    rounds it to 2 decimal places, and appends a '%' sign.

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to process
    - start_col: The starting column (default is 2 -> 'B')
    - end_col: The ending column (default is 53 -> 'BA')
    """

    for row_number in row_numbers:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_number, column=col)

            # Convert numeric strings (including negatives) to float
            if isinstance(cell.value, str) and cell.value.replace(".", "", 1).replace("-", "", 1).isdigit():
                cell.value = float(cell.value)

            # Ensure value is a number before modifying
            if isinstance(cell.value, (int, float)):
                cell.value = f"{round(cell.value * num, 2)}%"
                

def apply_dollar_format(ws, row_numbers, start_col=2, end_col=53):
    """
    Formats numeric values in specified rows:
    - Converts large numbers to dollar format ($1,000.00)
    - Converts small numbers (counts) to integers
    - Converts percentage values (e.g., '37.71%') to numeric format (0.3771) and applies percentage style

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to process
    - start_col: The starting column (default is 2 -> 'B')
    - end_col: The ending column (default is 53 -> 'BA')
    """

    for row_number in row_numbers:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_number, column=col)

            # # Convert percentage values
            # if isinstance(cell.value, str) and "%" in cell.value:
            #     try:
            #         percent_value = float(cell.value.replace("%", "")) / 100  # Convert to decimal
            #         cell.value = percent_value
            #         cell.number_format = '#.##%'  # Apply percentage format in Excel
            #     except ValueError:
            #         continue  # Skip if not a valid number

            # Convert numeric strings to float/int
            if isinstance(cell.value, str):
                cell_value = cell.value.replace(",", "").replace("$", "")
                if cell_value.replace(".", "", 1).replace("-", "", 1).isdigit():
                    cell.value = float(cell_value)  # Convert to number

            # Ensure value is numeric before modifying
            if isinstance(cell.value, (int, float)):
                if 0 <= cell.value < 999:  
                    cell.value = int(cell.value)
                elif cell.value >= 999:
                    cell.number_format = '$#,##0'  


def apply_bold_to_cells(ws, row_numbers, columns=None):
    """
    Applies bold formatting to specific rows and columns.

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to process
    - columns: A list of column numbers to apply bold formatting. If None, bold entire row.
    """
    for row_number in row_numbers:
        if columns:
            # Apply bold to specified columns in given rows
            for col in columns:
                cell = ws.cell(row=row_number, column=col)
                cell.font = Font(bold=True)
        else:
            # Apply bold to the entire row, skipping blank columns
            for col in range(2, ws.max_column + 1):  # Start from B (col 2)
                cell = ws.cell(row=row_number, column=col)
                if cell.value:  # Only format non-empty cells
                    cell.font = Font(bold=True)


def auto_adjust_column_width(ws):
    """
    Adjusts the width of all columns, setting a fixed width for the first column
    and adjusting the rest based on the data they contain.
    """
    # Set a fixed width for Column A (adjust the value as needed)
    ws.column_dimensions['A'].width = 5  # You can change 5 to any fixed width you prefer
    ws.column_dimensions['B'].width = 25

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        # Skip Column A since its width is already fixed
        if col_letter == 'A':
            continue

        # Set a fixed width (e.g., 15)
        ws.column_dimensions[col_letter].width = 15


def align(ws, row_numbers, format, start_col=2, end_col=53):
    """
    Aligns values to the right for specific rows and columns.

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to process
    - start_col: The starting column (default is 2 -> 'B')
    - end_col: The ending column (default is 100)
    """

    for row_number in row_numbers:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_number, column=col)
            cell.alignment = Alignment(horizontal=format)  # Set alignment to right
            
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Function to color specified rows
def apply_color_to_rows(ws, row_list):
    """Applies grey fill to given row numbers."""
    for row in row_list:
        for col in range(1, ws.max_column + 1):  # Loop through all columns
            ws.cell(row=row, column=col).fill = grey_fill  # Apply grey fill


def set_cell_value(ws, cell, value, font_size=12, bold=False):
    ws[cell] = value
    ws[cell].font = Font(size=font_size, bold=bold)


def delete_rows(ws, rows_to_delete):
    """
    Deletes specified rows from an openpyxl worksheet.
    
    Args:
        ws: openpyxl worksheet object
        rows_to_delete: List of row numbers to delete (1-based index)
    """
    for row in sorted(rows_to_delete, reverse=True):  
        ws.delete_rows(row)


def update_percentage_values(ws, row_numbers, col_names, both):
    """
    Removes the percent sign and multiplies the values by 100 for the given rows and columns in an openpyxl worksheet.

    :param ws: openpyxl worksheet object
    :param row_numbers: List of row numbers (1-based index)
    :param col_names: List of column names (Excel-style, e.g., 'A', 'B')
    """
    for row in row_numbers:
        for col_name in col_names:
            cell = ws[f"{col_name}{row}"]
            if isinstance(cell.value, str) and cell.value.endswith('%'):
                try:
                    # Convert to number, remove percent sign, multiply by 100
                    if both == True:
                        cell.value = float(cell.value.strip('%'))
                    else:
                        cell.value = f'{float(cell.value.strip("%")) * 100}%'
                except ValueError:
                    print(f"Skipping invalid value in {col_name}{row}: {cell.value}")
        


def merge_entire_row(ws, row_numbers, value_col=2):
    """
    Merges all columns in specific rows while keeping a value in Column B.

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to merge.
    - value_col: The column where the value should be placed (default is Column B).
    """
    for row in row_numbers:
        start_col = 1  # Start merging from Column A
        end_col = ws.max_column  # Merge up to the last column

        # Store the value from Column B before merging
        value = ws.cell(row=row, column=value_col).value  

        # Merge the entire row
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

        # Access the top-left cell of the merged region and assign the value
        merged_cell = ws.cell(row=row, column=start_col)
        merged_cell.value = value  # Assign the stored value
        merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def merge_and_border_preserve_text(ws, cell_ranges):
    """
    Merges given cell ranges, keeps the first non-empty text, centers it, 
    makes it bold, and applies a border around the merged area.

    Parameters:
    - ws: The worksheet object.
    - cell_ranges: A list of tuples containing (start_cell, end_cell) defining the range.
    """

    for start_cell, end_cell in cell_ranges:
        # Extract row numbers and convert column letters to indexes
        start_col, start_row = start_cell.rstrip('0123456789'), int(''.join(filter(str.isdigit, start_cell)))
        end_col, end_row = end_cell.rstrip('0123456789'), int(''.join(filter(str.isdigit, end_cell)))

        start_col_idx = column_index_from_string(start_col)  # Convert 'W' -> 23
        end_col_idx = column_index_from_string(end_col)      # Convert 'AF' -> 32

        # Find the first non-empty cell in the range
        text_to_keep = None
        for row in range(start_row, end_row + 1):
            for col in range(start_col_idx, end_col_idx + 1):  
                cell = ws[f"{get_column_letter(col)}{row}"]
                if cell.value:  # Keep the first non-empty value
                    text_to_keep = cell.value
                    break
            if text_to_keep:
                break

        # Merge the cells
        ws.merge_cells(start_row=start_row, start_column=start_col_idx,
                       end_row=end_row, end_column=end_col_idx)

        # Set the preserved text in the merged cell
        merged_cell = ws[f"{start_cell}"]
        merged_cell.value = text_to_keep

        # Apply alignment and bold font
        merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merged_cell.font = Font(bold=True)  # ✅ Make text bold


def weekly_credit_metrics_report(file_name, sdk, email_api, bucket, get_bucket):
    
    #risk score by % max upb
    response = sdk.run_look(str(308), "csv")
    data = pd.read_csv(io.StringIO(response))
    data = data.drop(data.columns[1], axis=1)
    risk_score_percenteage_upb = data.copy()

        #risk score by max upb
    response = sdk.run_look(str(309), "csv")
    data = pd.read_csv(io.StringIO(response))
    data = data.drop(data.columns[1], axis=1)
    risk_score_upb = data.copy()

     # other credit metrics
    response = sdk.run_look(str(310), "csv")
    cm_data = pd.read_csv(io.StringIO(response))
    cm_data = cm_data.drop(cm_data.columns[1], axis=1)
    columns_to_drop = cm_data.columns[range(1, 61, 2)]
    cm_data = cm_data.drop(columns=columns_to_drop, axis=1)
    cm_data = cm_data.copy()

    
    with pd.ExcelWriter('credit_metrics.xlsx', engine='openpyxl') as writer:
        risk_score_percenteage_upb.to_excel(writer, startrow=1, startcol=0, index = False)
        risk_score_upb.to_excel(writer, startrow=15, startcol=0, index = False)
        cm_data.to_excel(writer, startrow=30, startcol=0, index = False)


    ws_name = 'credit_metrics.xlsx'
    file_path = 'credit_metrics.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]
    ws.title = "Weekly Credit Metrics"
    delete_rows(ws, [2, 3, 5, 6, 16, 17, 19, 20, 31, 32, 34, 35])
    multiply_and_format(ws, row_numbers=[3, 4, 5, 6, 7, 8], num=100)
    apply_dollar_format(ws, row_numbers=[13, 14, 15, 16, 17, 18, 24, 25, 26, 27, 28, 29, 30])
    apply_bold_to_cells(ws, row_numbers=[2, 12, 23])
    # ws['F2'].font = Font(size=18, bold=True) 
    row_nums = [i for i in range(2, 65)]
    align(ws, row_numbers=row_nums, format="right")
    align(ws, row_numbers=[2, 12, 23], format="center")
    # Define specific rows to be colored
    rows_to_color = [4, 6, 8, 14, 16, 18, 25, 27, 29]
    # Apply the color
    apply_color_to_rows(ws, rows_to_color)
    auto_adjust_column_width(ws)
    # ws.freeze_panes = "C4"
    ws.insert_rows(1)
    ws.insert_rows(10)
    ws.insert_cols(1)

    regular_font = Font(size=10)  # 12pt normal font
    bold_large_font = Font(size=11, bold=True, color="1F497D")
    bold_large_category = Font(size=11, bold=True) 
    # Assign values with formatting
    ws['B1'] = "Allocation of Risk Scores by % of Purchased Max UPB"
    ws['B1'].font = bold_large_font  # Apply 11pt bold font
    ws['B10'] = "(a) Weighted-average by max UPB."
    ws['B10'].font = regular_font  # Apply 12pt font
    ws['B12'] = "Allocation of Risk Scores by Purchased Max UPB ($)"
    ws['B12'].font = bold_large_font  # Apply 11pt bold font
    ws['B23'] = "Other Credit Metrics"
    ws['B23'].font = bold_large_font  # Apply 11pt bold font
    
    ws['B32'] = "(b) % of Max UPB to experienced borrowers (3+)"
    ws['B32'].font = regular_font  # Apply 12pt font
    ws['B33'] = "(c) SFR 1-4 unit only. High Balance defined as Max UPB $750K+ ($1mm+ in CA). "
    ws['B33'].font = regular_font  # Apply 12pt font
    ws['B34'] = "(d) Purchase purpose only	"
    ws['B34'].font = regular_font  # Apply 12pt font
    
    ws['F2'] = "Prior Week"
    ws['F2'].font = bold_large_category  # Apply 12pt font
    ws['F13'] = "Prior Week"
    ws['F13'].font = bold_large_category  # Apply 12pt font
    ws['h24'] = "Prior Week"

    ws['M2'] = "Trailing 4 Weeks"
    ws['M2'].font = bold_large_category  # Apply 12pt font
    ws['M13'] = "Trailing 4 Weeks"
    ws['M13'].font = bold_large_category  # Apply 12pt font
    ws['R24'] = "Trailing 4 Weeks"
    ws['R24'].font = bold_large_category  # Apply 12pt font

    ws['T2'] = "1H2024"
    ws['T2'].font = bold_large_category  # Apply 12pt font
    ws['T13'] = "1H2024"
    ws['T13'].font = bold_large_category  # Apply 12pt font
    ws['AB24'] = "1H2024"
    ws['AB24'].font = bold_large_category  # Apply 12pt font

    merge_and_border_preserve_text(ws, [('C2', 'I2'), ('J2', 'P2'), ('Q2', 'W2'), ('C13', 'I13'), ('J13', 'P13'), ('Q13', 'W13'), ('C24', 'L24'), ('M24', 'V24'), ('W24', 'AF24')])
    wb.save(ws_name)
    set_cell_value(ws, 'B2', 'SFR 1-4 (no GUC)', font_size=12, bold=False)
    update_percentage_values(ws, [4,5,6,7,8,9], ['H', 'O', 'V'], True)
    update_percentage_values(ws, [4,5,6,7,8,9], ['I', 'P', 'W'], False)
    wb.save(ws_name)

    with open(file_path, 'rb') as f:
        txt=f.read()
        # print('read the file')
        blob = bucket.blob(f'looker_report_emails/{file_name}.xlsx')
        blob.upload_from_string(txt)
        attachments_ = txt

    text_ = 'this is test'

    body_ = """<html>
                <head>
                <ul id="ul-img" style="list-style-type: none;margin: 0;padding: 0;background-color: #0fcbef;width: 100%;height: 60px;border-radius: 4px;background-image: -webkit-linear-gradient(97deg, #0fcbef 4%, #1071ee 90%) !important;">
                    <div>
                    <img src="https://file-management-dev-tl-ue1-s3.s3.amazonaws.com/tooraklogo.png" />
                    </div>
                </ul>
            </head>
            <body>
            <br>Hi All,<br><br>Please find attached report name.<br>Please let us know if you face any issues.
            <br><br>Thanks,<br>Toorak.<br>
            <img src='https://static.wixstatic.com/media/74c4f9_20406c39aa61491d83b0ccec086c6feb~mv2_d_4500_4500_s_4_2.png/v1/fit/w_1000%2Ch_1000%2Cal_c/file.png' width="250" height="190">
            </body>
            </html>"""
    body_ = body_.replace('report name',file_name)

    response = send_mail(sender_email, cm_email_recipients, cm_email_cc, text_, body_, file_name, attachments_, file_name, email_api)