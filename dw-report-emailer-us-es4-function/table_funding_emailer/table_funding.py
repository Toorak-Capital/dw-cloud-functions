from table_funding_emailer.variables import * 
from variables import *
from send_email import *
import io
import re
from PIL import Image
import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
from google.cloud import storage
import json
from openpyxl.styles import Font
import jinja2

file_path = '/tmp/output.xlsx'

def apply_left_alignment(sheet):
    """Apply left alignment to all cells in the given sheet."""
    left_alignment = Alignment(horizontal='left')
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = left_alignment

def auto_adjust_columns(sheet):
    """
    Automatically adjust column width based on content.
    """
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width


def apply_currency_format(sheet, columns):
    """
    Apply currency formatting to specified columns in a worksheet.
    """
    dollar_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # "$#,##0.00"
    
    for col_letter in columns:
        for row in range(2, sheet.max_row + 1):  # Skip header row
            cell = sheet[f"{col_letter}{row}"]
            value = str(cell.value).strip() if cell.value is not None else None
            
            if value:
                # Remove any currency symbols, commas, or whitespace
                cleaned_value = re.sub(r"[^\d.-]", "", value)
                try:
                    # Convert to float and apply formatting
                    numeric_value = float(cleaned_value)
                    cell.value = numeric_value
                    cell.number_format = dollar_format
                except ValueError:
                    # Skip non-convertible values
                    pass

def table_funding_report(file_name, sdk, email_api, bucket, get_bucket):

    response = sdk.run_look(str(table_funding_look_ids['loan_info']), "csv") 
    loan_info = pd.read_csv(io.StringIO(response))
    loan_info_cols = ['Spoc','Broker','Toorak Loan ID','Originator Loan ID','Current Status','Date Converted to Initial Review','Date Converted to Review in Progress','Date Converted to Final Review','Date Converted to Approved','Property Address','Property Type','Loan Characterisation','Initial Loan Amount','Max Loan Amount','Inquiries','TC Fails (Y/N)',
'Risk Bucket']
    loan_info.columns = loan_info_cols
    gap = len(loan_info) + 5

    response = sdk.run_look(str(table_funding_look_ids['summary']), "csv")
    summary = pd.read_csv(io.StringIO(response))
    summary_cols = ['Loan Status','Initial Loan Amount','Maximum Loan Amount','Loan Count']
    summary.columns = summary_cols
    print(len(summary))
    
        
    with pd.ExcelWriter(file_path) as writer1:
        loan_info.to_excel(writer1, sheet_name="Loan_Information", startrow=0, startcol=0, index=False)
        summary.to_excel(writer1, sheet_name="Summary", startrow=0, startcol=0, index=False)

    # Load workbook and apply formatting
    wb = openpyxl.load_workbook(file_path)
    sheet1 = wb["Loan_Information"]
    sheet2 = wb["Summary"]

    # Auto-adjust column widths
    auto_adjust_columns(sheet1)
    auto_adjust_columns(sheet2)

    # Apply formatting and alignment
    apply_left_alignment(sheet1)
    apply_left_alignment(sheet2)
    apply_currency_format(sheet1, ["M", "N"])  # Loan Information sheet: Columns M, N
    apply_currency_format(sheet2, ["B", "C"]) 

    # Save workbook
    wb.save(file_path)

    with open(file_path, 'rb') as f:
        txt=f.read()
        blob = bucket.blob(f'looker_report_emails/{file_name}.xlsx')
        blob.upload_from_string(txt)
        attachments_ = txt


    title_ = file_name
    text_ = ''
    
    
    recipients_ = table_funding_recipients
    cc_ = table_funding_cc

    body_ = """<html>
                <head>
                <ul id="ul-img" style="list-style-type: none;margin: 0;padding: 0;background-color: #0fcbef;width: 100%;height: 60px;border-radius: 4px;background-image: -webkit-linear-gradient(97deg, #0fcbef 4%, #1071ee 90%) !important;">
                    <div>
                    <img src="https://file-management-dev-tl-ue1-s3.s3.amazonaws.com/tooraklogo.png" />
                    </div>
                </ul>
            </head>
            <body>
            <br>Hi All,<br><br>Please find attached report name.<br>Please let us know if you face any issues.<br><br>Thanks,<br>Toorak.<br>
            <img src='https://static.wixstatic.com/media/74c4f9_20406c39aa61491d83b0ccec086c6feb~mv2_d_4500_4500_s_4_2.png/v1/fit/w_1000%2Ch_1000%2Cal_c/file.png' width="250" height="190">
            </body>
            </html>"""
    body_ = body_.replace('report name',file_name)
        
    response = send_mail(sender_email, table_funding_recipients, table_funding_cc, text_, body_, file_name, attachments_, file_name, email_api)
 