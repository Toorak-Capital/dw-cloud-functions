from variables import * 
from send_email import *
import io
from PIL import Image
import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from google.cloud import storage
import json
from openpyxl.styles import Font
import jinja2
from merchants_emailer.variables import *

def modify_excel_cells(ws, rows_to_modify):
    """
    Modify specified rows in the Excel worksheet.

    Parameters:
    ws (Worksheet): The worksheet to modify.
    rows_to_modify (list): List of row numbers to modify.
    """
    for row in rows_to_modify:
        for col in range(2, 54):  # B is column 2 and BA is column 53
            cell = ws.cell(row=row, column=col)
            
            if isinstance(cell.value, str):  # Ensure the cell contains a string
                if row in [4, 20, 30, 40]:
                    new_value = cell.value.split('.')[0]  # Keep the part before '.'
                    cell.value = new_value
                
                # For row 24: Append percentage sign
                if row in [24, 35, 45]:
                    cell.value = f"{cell.value}%"

# Define USD currency format
    usd_format = '#,##0.00'
 
#     # Apply USD format to the range B8:BA16
#     for row in ws.iter_rows(min_row=5, max_row=35, min_col=2, max_col=53):  # B=2, BA=53
#         for cell in row:
#             cell.number_format = usd_format
            
    specific_rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 34, 44]

    # Apply USD format to specific rows and columns (B to BA)
    for row in ws.iter_rows(min_row=5, max_row=50, min_col=2, max_col=53):
        for cell in row:
            if cell.row in specific_rows:  # Check if the current row is in the list
                cell.number_format = usd_format

def merchant_weekly_report(file_name, sdk, email_api, bucket, get_bucket):

    #first report
    response = sdk.run_look(str(merchant_look_ids['purchases_1']), "csv")
    data = pd.read_csv(io.StringIO(response))
    data = data.drop(data.columns[0], axis=1)
    data = data.drop(data.index[-1])
    data = data.drop(data.index[3])
    rotated_df = data.T[::1]
    rotated_df.columns = rotated_df.iloc[0]  
    rotated_df = rotated_df[1:] 
    # rotated_df.reset_index(drop=True, inplace=True)
    rotated_df.fillna(0, inplace=True)
    columns_to_convert = ['CA', 'CO', 'AZ', 'WA', 'OR', 'TX', 'MO', 'Non-Top 7']
    for column in columns_to_convert:
        rotated_df[column] = pd.to_numeric(rotated_df[column], errors='coerce')
    rotated_df['Total Sum'] = rotated_df['CA'] + rotated_df['CO'] + rotated_df['AZ'] + rotated_df['WA'] + rotated_df['OR'] + rotated_df['TX'] + rotated_df['MO'] + rotated_df['Non-Top 7']
    rotated_df_1 = rotated_df.T[::1]
    index_mapping = {
        'Merchant Weekly Purchases Table Row Num': 'Week Number',
        'Merchant Weekly Purchases Table Week Range': 'Week Range',
        'Merchant Weekly Purchases Table Month': 'Month'
    }
    
    # Rename indices in one go
    rotated_df_1 = rotated_df_1.rename(index=index_mapping)
    
    #first table part 2
    response = sdk.run_look(str(merchant_look_ids['purchases_2']), "csv")
    data = pd.read_csv(io.StringIO(response))
    data = data.drop(data.columns[0], axis=1)
    data = data.drop(data.index[3])
    rotated_df = data.T[::1]
    rotated_df.columns = rotated_df.iloc[0]  # Set the first row as header
    rotated_df = rotated_df[1:]               # Remove the first row from the DataFrame
    # rotated_df.reset_index(drop=True, inplace=True)
    rotated_df.fillna(0, inplace=True)
    columns_to_convert = ['1', '2', '3', '4', '5', '6', '7', '8']
    for column in columns_to_convert:
        rotated_df[column] = pd.to_numeric(rotated_df[column], errors='coerce')
    rotated_df['Total Count'] = rotated_df['1'] + rotated_df['2'] + rotated_df['3'] + rotated_df['4'] + rotated_df['5'] + rotated_df['6'] + rotated_df['7'] + rotated_df['8']
    rotated_df = rotated_df.T[::1]
    rotated_df = rotated_df.tail(1)
    rotated_df_1 = pd.concat([rotated_df_1, rotated_df])


    #rolling 4
    response = sdk.run_look(str(merchant_look_ids['rolling_4_1']), "csv")
    data_1 = pd.read_csv(io.StringIO(response))
    data_1 = data_1.drop(data_1.index[3])

    response = sdk.run_look(str(merchant_look_ids['rolling_4_2']), "csv")
    data_2 = pd.read_csv(io.StringIO(response)).tail(1)
    data_1 = pd.concat([data_1, data_2])

    response = sdk.run_look(str(merchant_look_ids['rolling_4_3']), "csv")
    data_3 = pd.read_csv(io.StringIO(response)).tail(1)
    data_final = pd.concat([data_1, data_3])

    response = sdk.run_look(str(merchant_look_ids['wac']), "csv")
    data = pd.read_csv(io.StringIO(response))
    data = data.drop(data.index[3])
    data = data.rename(columns={'Merchant Weekly Wac Table Current Year': 'WAC'})
    replacements = {
        'Merchant Weekly Wac Table Row Num': 'Week Number',
        'Merchant Weekly Wac Table Week Range': 'Week Range',
        'Merchant Weekly Wac Table Month': 'Month'
    }

    data = data.replace(replacements)
    

    response = sdk.run_look(str(merchant_look_ids['rolling_8_1']), "csv")
    data_8_1 = pd.read_csv(io.StringIO(response))
    data_8_1 = data_8_1.drop(data_8_1.index[3])

    response = sdk.run_look(str(merchant_look_ids['rolling_8_2']), "csv")
    data_8_2 = pd.read_csv(io.StringIO(response)).tail(1)
    data_8_1 = pd.concat([data_8_1, data_8_2])
    data_8_1 = data_8_1.rename(columns={'Merchant Weekly Rolling 8 Table Current Year': 'Rolling 8'})
    replacements_8_1 = {
        'Merchant Weekly Rolling 8 Table Row Num': 'Week Number',
        'Merchant Weekly Rolling 8 Table Week Range': 'Week Range',
        'Merchant Weekly Rolling 8 Table Month': 'Month'
    }

    data_8_1 = data_8_1.replace(replacements_8_1)

    with pd.ExcelWriter('/tmp/output.xlsx', engine='openpyxl') as writer:
        rotated_df_1.to_excel(writer, startrow=3, startcol=0)
        data.to_excel(writer, startrow=19, startcol=0, index = False)
        data_final.to_excel(writer, startrow=29, startcol=0, index = False)
        data_8_1.to_excel(writer, startrow=39, startcol=0, index = False)

    
    #first graph
    response = sdk.run_look(str(merchant_look_ids['graph_1']), "png", image_width=1500, image_height=500)
    image = Image.open(io.BytesIO(response))
    image.save('view.png')
    ws_name = '/tmp/output.xlsx'
    file_path = '/tmp/output.xlsx'
    wb = openpyxl.load_workbook(file_path)
    wb.create_sheet("Data Visualisation",1)
    ws = wb["Data Visualisation"]
    img = openpyxl.drawing.image.Image('view.png')
    img.anchor = 'B6'
    img.width = 850
    ws.add_image(img)

    #second graph
    response = sdk.run_look(str(merchant_look_ids['graph_2']), "png", image_width=1500, image_height=500)
    image = Image.open(io.BytesIO(response))
    image.save('view_1.png')
    img = openpyxl.drawing.image.Image('view_1.png')
    img.anchor = 'B40'
    img.width = 850
    ws.add_image(img)

    #third graph
    response = sdk.run_look(str(merchant_look_ids['graph_3']), "png", image_width=1500, image_height=500)
    image = Image.open(io.BytesIO(response))
    image.save('view_2.png')
    img = openpyxl.drawing.image.Image('view_2.png')
    img.anchor = 'B68'
    img.width = 850
    ws.add_image(img)
    ws['F2'] = "WAC Bridge"  
    ws['F2'].font = Font(size=18, bold=True) 
    ws['F36'] = "Origination Volume"  
    ws['F36'].font = Font(size=18, bold=True) 
    ws['F65'] = "Purchase By State"  
    ws['F65'].font = Font(size=18, bold=True) 
    ws = wb["Sheet1"]
    ws.freeze_panes = "B1" 
    ws['A4'] = "Purchases ($) (max balance)"
    rows_to_modify = [4, 20, 30, 40, 24, 35, 45]
    modify_excel_cells(ws, rows_to_modify)
    wb.save(ws_name)

    with open(file_path, 'rb') as f:
        txt=f.read()
        # print('read the file')
        blob = bucket.blob(f'looker_report_emails/{file_name}.xlsx')
        blob.upload_from_string(txt)
        attachments_ = txt
        
    text_ = 'this is test'
    body_ = """<html>
                <head>
                <ul id="ul-img" style="list-style-type: none;margin: 0;padding: 0;background-color: #0fcbef;width: 100%;height: 60px;border-radius: 4px;background-image: -webkit-linear-gradient(97deg, #0fcbef 4%, #1071ee 90%) !important;">
                    <div>
                    <img src="https://file-management-dev-tl-ue1-s3.s3.amazonaws.com/tooraklogo.png" />
                    </div>
                </ul>
            </head>
            <body>
            <br>Hi All,<br><br>Please find attached report name.<br>Please let us know if you face any issues.<br><br>Thanks,<br>Toorak.<br>
            <img src='https://static.wixstatic.com/media/74c4f9_20406c39aa61491d83b0ccec086c6feb~mv2_d_4500_4500_s_4_2.png/v1/fit/w_1000%2Ch_1000%2Cal_c/file.png' width="250" height="190">
            </body>
            </html>"""
    body_ = body_.replace('report name',file_name)
        
    response = send_mail(sender_email, merchant_email_recipients, merchant_email_cc, text_, body_, file_name, attachments_, file_name, email_api)
    
 