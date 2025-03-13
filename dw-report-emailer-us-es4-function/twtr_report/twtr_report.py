from variables import * 
from send_email import *
import io
import os
from PIL import Image
import os
import pandas as pd
from datetime import *
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from google.cloud import storage
import json
from openpyxl.styles import Font
import jinja2
from twtr_report.variables import *


def move_rows(source_range, target_range, ws, clear):
    """
    Moves data from source rows to target rows.
    Clears the original source rows after copying.
    """
    for source, target in zip(source_range, target_range):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=target, column=col, value=ws.cell(row=source, column=col).value)

    # Clear original source rows
    if clear == 1:
        for row in source_range:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col, value="")

def auto_adjust_column_width(ws):
    """
    Adjusts the width of all columns, setting a fixed width for the first column
    and adjusting the rest based on the data they contain.
    """
    # Set a fixed width for Column A (adjust the value as needed)
    ws.column_dimensions['A'].width = 25  # You can change 25 to any fixed width you prefer

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        # Skip Column A since its width is already fixed
        if col_letter == 'A':
            continue

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        ws.column_dimensions[col_letter].width = max_length + 2



def modify_excel_cells(ws, rows_to_modify):
    for row in rows_to_modify:
        for col in range(2, 54):  # B is column 2 and BA is column 53
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):  # Ensure the cell contains a string
                # Remove the '.x' part if it exists
                new_value = cell.value.split('.')[0]  # Keep the part before '.'
                cell.value = new_value




def multiply_and_format(ws, row_numbers, num, start_col=2, end_col=53):
    """
    Multiplies each cell's value in the specified rows by 100,
    rounds it to 2 decimal places, and appends a '%' sign.

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to process
    - start_col: The starting column (default is 2 -> 'B')
    - end_col: The ending column (default is 53 -> 'BA')
    """

    for row_number in row_numbers:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_number, column=col)

            # Convert numeric strings (including negatives) to float
            if isinstance(cell.value, str) and cell.value.replace(".", "", 1).replace("-", "", 1).isdigit():
                cell.value = float(cell.value)

            # Ensure value is a number before modifying
            if isinstance(cell.value, (int, float)):
                cell.value = f"{round(cell.value * num, 2)}%"




def apply_dollar_format(ws, row_numbers, start_col=2, end_col=53):
    """
    Converts each numeric cell's value in the specified rows to dollar format ($),
    rounding to 2 decimal places.

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to process
    - start_col: The starting column (default is 2 -> 'B')
    - end_col: The ending column (default is 53 -> 'BA')
    """

    for row_number in row_numbers:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_number, column=col)

            # Convert numeric strings (including negatives) to float
            if isinstance(cell.value, str) and cell.value.replace(".", "", 1).replace("-", "", 1).isdigit():
                cell.value = float(cell.value)

            # Ensure value is a number before modifying
            if isinstance(cell.value, (int, float)):
                cell.value = f"${round(cell.value, 2):,.2f}"  # Format as dollars with commas


def apply_bold_to_cells(ws, row_numbers, columns):
    """
    Applies bold formatting to specific cells in given rows and columns.

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to process
    - columns: A list of column numbers to apply bold formatting
    """

    for row_number in row_numbers:
        for col in columns:
            cell = ws.cell(row=row_number, column=col)
            cell.font = Font(bold=True)  # Apply bold formatting


def insert_image(sdk, ws, look_id, file_name, anchor_cell, width=700, height=300):
    """
    Inserts an image into the given worksheet at a specified location.

    Parameters:
    - ws: The worksheet object.
    - look_id: The Looker API ID for fetching the image.
    - file_name: The name to save the image as (e.g., 'view_10.png').
    - anchor_cell: The Excel cell where the image should be placed.
    - width: Image width (default: 700).
    - height: Image height (default: 300).
    """

    # Get image from Looker
    response = sdk.run_look(str(look_id), "png", image_width=width, image_height=height)
    image = Image.open(io.BytesIO(response))
    image.save(file_name)

    # Insert into worksheet
    img = openpyxl.drawing.image.Image('view_10.png')
    img.anchor = anchor_cell
    img.width = width
    ws.add_image(img)



def align(ws, row_numbers, format, start_col=2, end_col=53):
    """
    Aligns values to the right for specific rows and columns.

    Parameters:
    - ws: The worksheet object
    - row_numbers: A list of row numbers to process
    - start_col: The starting column (default is 2 -> 'B')
    - end_col: The ending column (default is 53 -> 'BA')
    """

    for row_number in row_numbers:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_number, column=col)
            cell.alignment = Alignment(horizontal=format)  # Set alignment to right

def set_cell_value(ws, cell, value, font_size=12, bold=False):
    ws[cell] = value
    ws[cell].font = Font(size=font_size, bold=bold)



def twtr_report(file_name, sdk, email_api, bucket, get_bucket):
    
    #first report
    response = sdk.run_look(str(290), "csv")
    data = pd.read_csv(io.StringIO(response))
    columns_to_drop = data.columns[range(3, 107, 2)]
    data = data.drop(columns=columns_to_drop, axis=1)
    data = data.drop(data.columns[2], axis=1)
    purchase_data = data.drop(data.columns[0], axis=1)

    
    # non ga count
    response = sdk.run_look(str(291), "csv")
    ga_count = pd.read_csv(io.StringIO(response))
    ga_count = ga_count.drop(ga_count.columns[0], axis=1)
    ga_count = ga_count.iloc[3:]

    # payoffs data
    response = sdk.run_look(str(293), "csv")
    payoffs_data = pd.read_csv(io.StringIO(response))

     # wac data
    response = sdk.run_look(str(289), "csv")
    wac_data = pd.read_csv(io.StringIO(response))
    wac_data = wac_data.drop(wac_data.columns[1], axis=1)

     # net-wac data
    response = sdk.run_look(str(294), "csv")
    net_wac_data = pd.read_csv(io.StringIO(response))
    net_wac_data = net_wac_data.drop(net_wac_data.columns[1], axis=1)

     # first loss data
    response = sdk.run_look(str(295), "csv")
    first_loss_data = pd.read_csv(io.StringIO(response))
    # first_loss_data = first_loss_data.drop(first_loss_data.columns[1], axis=1)

     # pp_data
    response = sdk.run_look(str(296), "csv")
    pp_data = pd.read_csv(io.StringIO(response))
    # first_loss_data = first_loss_data.drop(first_loss_data.columns[1], axis=1)

     # cm_data
    response = sdk.run_look(str(297), "csv")
    cm_data = pd.read_csv(io.StringIO(response))
    columns_to_drop = cm_data.columns[range(3, 106, 2)]
    cm_data = cm_data.drop(columns=columns_to_drop, axis=1)
    cm_data = cm_data.drop(cm_data.columns[0], axis=1)
    cm_data = cm_data.drop(cm_data.columns[1], axis=1)

     # r4_data
    response = sdk.run_look(str(300), "csv")
    r4_data = pd.read_csv(io.StringIO(response))
    columns_to_drop = r4_data.columns[range(2, 106, 2)]
    r4_data = r4_data.drop(columns=columns_to_drop, axis=1)
    r4_data = r4_data.drop(r4_data.columns[1], axis=1)

    # weighted_data
    response = sdk.run_look(str(283), "csv")
    weighted_data = pd.read_csv(io.StringIO(response))
    columns_to_drop = weighted_data.columns[range(2, 106, 2)]
    weighted_data = weighted_data.drop(columns=columns_to_drop, axis=1)
    weighted_data = weighted_data.drop(weighted_data.columns[1], axis=1)

     # week_delta_data
    response = sdk.run_look(str(301), "csv")
    week_delta_data = pd.read_csv(io.StringIO(response))
    week_delta_data = week_delta_data.drop(week_delta_data.columns[1], axis=1)
    
    # yoy
    response = sdk.run_look(str(270), "csv")
    yoy_data = pd.read_csv(io.StringIO(response))
    columns_to_drop = yoy_data.columns[range(2, 106, 2)]
    yoy_data = yoy_data.drop(columns=columns_to_drop, axis=1)
    yoy_data = yoy_data.drop(yoy_data.columns[1], axis=1)

    # payoff_count
    response = sdk.run_look(str(302), "csv")
    payoff_count = pd.read_csv(io.StringIO(response))
    
    with pd.ExcelWriter('/tmp/output.xlsx', engine='openpyxl') as writer:
        purchase_data.to_excel(writer, startrow=1, startcol=0, index = False)
        ga_count.to_excel(writer, startrow=22, startcol=0, index = False)
        payoffs_data.to_excel(writer, startrow=29, startcol=0, index = False)
        wac_data.to_excel(writer, startrow=39, startcol=0, index = False)
        net_wac_data.to_excel(writer, startrow=51, startcol=0, index = False)
        first_loss_data.to_excel(writer, startrow=64, startcol=0, index = False)
        pp_data.to_excel(writer, startrow=73, startcol=0, index = False)
        cm_data.to_excel(writer, startrow=85, startcol=0, index = False)
        r4_data.to_excel(writer, startrow=97, startcol=0, index = False)
        weighted_data.to_excel(writer, startrow=109, startcol=0, index = False)
        week_delta_data.to_excel(writer, startrow=122, startcol=0, index = False)
        yoy_data.to_excel(writer, startrow=222, startcol=0, index = False)
        payoff_count.to_excel(writer, startrow=322, startcol=0, index = False)


    ws_name = '/tmp/output.xlsx'
    file_path = '/tmp/output.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]
    ws['A3'] = "US Month"
    ws.delete_rows(4, 2)
    ws.insert_rows(4)
    ws.insert_rows(5)
    ws['A2'] = ''
    ws['A4'] = 'Purchases & Draws ($)'
    ws['A5'] = 'Non-GA'
    move_rows(range(15, 18), range(169, 172), ws, 1)
    move_rows(range(24, 25), range(15, 16), ws, 1)
    move_rows(range(34, 36), range(16, 18), ws, 1)
    ws.delete_rows(20, 24)
    ws['A19'] = 'WAC'
    ws.delete_rows(28, 4)
    ws.insert_rows(27)
    ws['A28'] = 'NET-WAC'
    move_rows(range(104, 105), range(35, 36), ws, 1)
    ws.delete_rows(38, 4)
    ws.insert_rows(37)
    ws['A38'] = 'First Loss'
    ws.delete_rows(43, 5)
    ws['A42'] = 'Purchase Premium'
    ws.delete_rows(48, 7)
    ws['A47'] = 'Credit Metrics (Bridge 1-4 Only)'
    ws.delete_rows(56, 4)
    move_rows(range(60, 62), range(90, 92), ws, 1)
    ws.delete_rows(64)
    move_rows(range(67, 75), range(60, 68), ws, 0)
    ws['A55'] = 'Rolling 4'
    move_rows(range(76, 79), range(76, 79), ws, 1)
    ws.delete_rows(76)
    move_rows(range(88, 90), range(75, 77), ws, 1)
    move_rows(range(79, 83), range(84, 89), ws, 1)
    ws.delete_rows(78)
    move_rows(range(120, 122), range(77, 79), ws, 1)
    ws['A82'] = '4 week avg Delta'
    move_rows(range(122, 123), range(87, 88), ws, 1)
    ws.delete_rows(80)
    ws.delete_rows(45)
    ws.delete_rows(67, 7)
    move_rows(range(168, 172), range(82, 86), ws, 1)
    move_rows(range(2, 3), range(81, 82), ws, 0)
    ws['A81'] = 'YOY'
    ws.delete_rows(165, 5)
    rows_to_modify = [2]
    modify_excel_cells(ws, rows_to_modify)
    multiply_and_format(ws, row_numbers=[20, 21, 22, 23, 24, 25, 29, 30, 31, 32, 33, 34, 35, 39, 43 ,74, 75, 76, 77], num=100)
    apply_dollar_format(ws, row_numbers=[6, 7, 8, 9, 10, 11, 12, 13, 14, 16, 17, 47, 55 ,56, 57, 58 ,61, 82, 84])
    apply_bold_to_cells(ws, row_numbers=[3, 4, 19, 28, 38, 42, 46, 54, 73], columns=[1])
    apply_bold_to_cells(ws, row_numbers=[81], columns=[i for i in range(1, 53)])
    row_nums = [i for i in range(4, 86)]
    align(ws, row_numbers=row_nums, format="right")
    align(ws, row_numbers=[3], format="center")
    auto_adjust_column_width(ws)
    ws.freeze_panes = "B4" 
    ws.insert_rows(69)
    ws.insert_rows(18)
    move_rows(range(266, 267), range(18, 19), ws, 1)
    ws.delete_rows(262, 7)
    align(ws, row_numbers=[18], format="right")
    ws.insert_rows(19)
    wb.save(ws_name)
    
    response = sdk.run_look(str(264), "png", image_width=700, image_height=300)
    image = Image.open(io.BytesIO(response))
    image.save('view_1.png')
    ws_name = '/tmp/output.xlsx'
    file_path = '/tmp/output.xlsx'
    wb = openpyxl.load_workbook(file_path)
    wb.create_sheet("Trended Charts",1)
    ws = wb["Trended Charts"]
    img = openpyxl.drawing.image.Image('view_1.png')
    img.anchor = 'B5'
    img.width = 700
    ws.add_image(img)

    insert_image(sdk, ws, look_id=272, file_name="view_2.png", anchor_cell="M5")
    insert_image(sdk, ws, look_id=281, file_name="view_3.png", anchor_cell="X5")
    insert_image(sdk, ws, look_id=279, file_name="view_4.png", anchor_cell="AI5")
    insert_image(sdk, ws, look_id=268, file_name="view_5.png", anchor_cell="AT5")
    insert_image(sdk, ws, look_id=282, file_name="view_6.png", anchor_cell="B30")
    insert_image(sdk, ws, look_id=280, file_name="view_7.png", anchor_cell="M30")
    insert_image(sdk, ws, look_id=275, file_name="view_8.png", anchor_cell="X30")
    insert_image(sdk, ws, look_id=278, file_name="view_9.png", anchor_cell="B55")
    insert_image(sdk, ws, look_id=277, file_name="view_10.png", anchor_cell="M55")
    insert_image(sdk, ws, look_id=276, file_name="view_11.png", anchor_cell="X55")
    insert_image(sdk, ws, look_id=284, file_name="view_12.png", anchor_cell="AI55")
    insert_image(sdk, ws, look_id=266, file_name="view_13.png", anchor_cell="B80")
    insert_image(sdk, ws, look_id=267, file_name="view_14.png", anchor_cell="M80")
    insert_image(sdk, ws, look_id=265, file_name="view_15.png", anchor_cell="X80")
 
    
    set_cell_value(ws, 'E3', 'Purchase Volume', font_size=18, bold=True)
    set_cell_value(ws, 'E28', 'Payoffs - Bridge', font_size=18, bold=True)
    set_cell_value(ws, 'E53', 'FICO', font_size=18, bold=True)
    set_cell_value(ws, 'E78', '% Merchant', font_size=18, bold=True)

    set_cell_value(ws, 'P3', 'Bridge Volume', font_size=18, bold=True)
    set_cell_value(ws, 'P28', 'Average Balance', font_size=18, bold=True)
    set_cell_value(ws, 'P53', 'As Is LTV', font_size=18, bold=True)
    set_cell_value(ws, 'P78', '% MF', font_size=18, bold=True)


    set_cell_value(ws, 'AA3', 'Draws', font_size=18, bold=True)
    set_cell_value(ws, 'AA28', '% Purchase', font_size=18, bold=True)
    set_cell_value(ws, 'AA53', 'ARLTV', font_size=18, bold=True)
    set_cell_value(ws, 'AA78', '% GU', font_size=18, bold=True)
    

    set_cell_value(ws, 'AL3', 'WAC - Bridge', font_size=18, bold=True)
    set_cell_value(ws, 'AL53', 'Originator Spread (ex Merchants)', font_size=18, bold=True)

    set_cell_value(ws, 'AW3', 'Payoff Volume', font_size=18, bold=True)
    
    wb.save(ws_name)


    with open(file_path, 'rb') as f:
        txt=f.read()
        # print('read the file')
        blob = bucket.blob(f'looker_report_emails/{file_name}.xlsx')
        blob.upload_from_string(txt)
        attachments_ = txt
        
    text_ = 'this is test'
    body_ = """<html>
                <head>
                <ul id="ul-img" style="list-style-type: none;margin: 0;padding: 0;background-color: #0fcbef;width: 100%;height: 60px;border-radius: 4px;background-image: -webkit-linear-gradient(97deg, #0fcbef 4%, #1071ee 90%) !important;">
                    <div>
                    <img src="https://file-management-dev-tl-ue1-s3.s3.amazonaws.com/tooraklogo.png" />
                    </div>
                </ul>
            </head>
            <body>
            <br>Hi All,<br><br>Please find attached report name.<br>Please let us know if you face any issues.<br><br>Thanks,<br>Toorak.<br>
            <img src='https://static.wixstatic.com/media/74c4f9_20406c39aa61491d83b0ccec086c6feb~mv2_d_4500_4500_s_4_2.png/v1/fit/w_1000%2Ch_1000%2Cal_c/file.png' width="250" height="190">
            </body>
            </html>"""
    body_ = body_.replace('report name',file_name)
        
    response = send_mail(sender_email, twtr_email_recipients, twtr_email_cc, text_, body_, file_name, attachments_, file_name, email_api)
    
 