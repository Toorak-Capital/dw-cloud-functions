import functions_framework
import looker_sdk
import os
import pandas as pd
import numpy as np
import json
import io
import boxsdk
from json import loads
from boxsdk import JWTAuth,Client
import os
import re
from datetime import datetime
import time
from google.cloud import bigquery
from google.cloud import storage
from variables import *
from google.cloud import secretmanager_v1
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
# Imports the Cloud Logging client library
import google.cloud.logging
from PIL import Image
# Instantiates a client
client = google.cloud.logging.Client()
client.setup_logging()
import logging

todays_date = datetime.now().strftime("%m-%d-%Y")
wells_file_name = f'DW - Wells IPS {todays_date}.xlsx'
wells_comparator_file_name = f'DW - Wells Comparator {todays_date}.xlsx'
pst_comparator_file_name = f'DW - PST Comparator {todays_date}.xlsx'
pst_file_name = f'DW - Payment Status Tracker {todays_date}.xlsx'
opr_file_name = f'DW - Originator Performance Report {todays_date}.xlsx'

def retriveBoxConfigFromSecret(secret):

    privateKey = secret['privateKey'].replace("\\n","\n")

    json_object = {
      "boxAppSettings": {
        "clientID": secret['clientID'] ,
        "clientSecret": secret['clientSecret'],
        "appAuth": {
          "publicKeyID": secret['publicKeyID'] ,
          "privateKey": privateKey,
          "passphrase": secret['passphrase']
        }
      },
      "enterpriseID": secret['enterpriseID']
    }
    box_user_id = secret['box_user_id']

    return json_object,box_user_id


def query_bigquery():
    client = bigquery.Client()
    result = False

    query = """
        SELECT * FROM reporting.dw_pipeline_log WHERE run_finished_time >= DATETIME(CONCAT(CAST(CURRENT_DATE("Asia/Calcutta") AS STRING), ' 17:30:00')) LIMIT 1
    """

    query_job = client.query(query)

    results = query_job.result()  # Waits for the job to complete
    for row in results:
        date = row['run_finished_time'].date()
        logging.info(date)
        if date == datetime.now().date():
            result = True
    return result

def check_log_file_in_gcs(bucket_name):
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    
    # List blobs in the bucket
    blobs = bucket.list_blobs()
    
    # Check for any .log files
    for blob in blobs:
        if blob.name.endswith('.log'):
            return True  # At least one .log file exists
    
    return False  # No .log files found


def get_secret(secret_id):

    client = secretmanager_v1.SecretManagerServiceClient()
    name = f"projects/{secret_project_id}/secrets/{secret_id}/versions/latest"
    response = client.access_secret_version(name=name)
    secret_data = response.payload.data.decode("UTF-8")
    try:
        secret_dict = json.loads(secret_data)
        return secret_dict
    except json.JSONDecodeError:
        # If it's not JSON, return the data as a plain string
        return secret_data



@functions_framework.http
def check_pipeline_run(request):

    log_file_exists = check_log_file_in_gcs(log_bucket_name)
    pipeline_ran_today = query_bigquery()

    if log_file_exists or not pipeline_ran_today:

        logging.info('Either files are missing or Pipeline did not ran. Cannot send emails')
        return {
        'statusCode': 500,
        'body': json.dumps('Either files are missing or Pipeline did not ran. Cannot send emails')
        }
    else:

        box_looker_conn()

        return {
            'statusCode': 200,
            'body': json.dumps({"message": "uploaded all files into box"})
        }

def dollar_sign(df,float_columns):
    
    # Add $ sign to float columns
    
    for col in float_columns:
        if col in list(df.columns):
            logging.info(col)
            df[col] = df[col].astype(float)
            df[col] = df[col].apply(lambda x: f'${x:,.2f}' if pd.notna(x) else np.nan)
    return df
    
def percentage_sign(df,float_columns):
    
    # Add $ sign to float columns
    for col in float_columns:
        if col in list(df.columns):
            logging.info(col)
            df[col] = df[col].astype(float)
            df[col] = df[col].apply(lambda x: f'{x*100:,.2f}%' if pd.notna(x) else np.nan)
    return df

def date_column_format(df,columns):

    for col in columns:
        if col in df.columns:
            logging.info(col)
    
            def safe_to_datetime(date_str):
                try:
                    
                    return pd.to_datetime(date_str).date()
                except pd.errors.OutOfBoundsDatetime:
                    
                    return date_str
                except Exception as e:
                    
                    logging.info(f"Error parsing date '{date_str}': {e}")
                    return date_str
    
            df[col] = df[col].apply(safe_to_datetime)
    
    return df
    
def run_look_and_clean_df(sdk, look_id, col_name):
    
    row_limit = 7000
    response = sdk.run_look(look_id, "csv", limit = row_limit)
    df = pd.read_csv(io.StringIO(response), dtype=str)
    df.columns = [col.replace('_', ' ') if col_name not in col else col.replace(f'{col_name} ','').strip() for col in df.columns]
    
    
    if 'mba order' in df.columns:
        df.drop('mba order', axis = 1, inplace = True)

    df = date_column_format(df, date_columns)
    df = dollar_sign(df, dollar_columns)
    df = percentage_sign(df, percentage_columns)
    
    return df

def run_look_and_save_png(sdk, look_id, file_path):
     
    response = sdk.run_look(look_id, "png")
    image = Image.open(io.BytesIO(response))
    image.save(file_path)

def pst_file_prep(sdk, user_client, pst_look_ids, pst_comparator_look_id, pst_comparator_summary_look_id):


    logging.info('inside pst prep')
    tmp_file = '/tmp/report.xlsx'
    pst_all_loans = run_look_and_clean_df(sdk, pst_look_ids['pst_all_loans'],'Payment Status Tracker')
    all_loans_summary = run_look_and_clean_df(sdk, pst_look_ids['pst_summary'],'Pst Summary')
    bridge_summary = run_look_and_clean_df(sdk, pst_look_ids['pst_bridge_summary'], 'Pst Bridge Summary')
    dscr_summary = run_look_and_clean_df(sdk, pst_look_ids['pst_dscr_summary'],'Pst Dscr Summary')

    empty_column = pd.DataFrame({ ' ': [np.nan] * len(bridge_summary) })

    pst_summary = pd.concat([all_loans_summary,empty_column,empty_column,bridge_summary,empty_column,empty_column,dscr_summary], axis=1)

    with pd.ExcelWriter(tmp_file, engine='openpyxl') as writer:
        pst_summary.to_excel(writer, index=False, header=True, sheet_name='Summary', startrow=1)
        pst_all_loans.to_excel(writer, index=False, header=True, sheet_name='PST')

    wb = load_workbook(tmp_file)
    ws = wb['Summary']

    ws.merge_cells('A1:D1')
    ws['A1'] = 'All Loans'
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


    ws.merge_cells('G1:J1')
    ws['G1'] = 'Bridge'
    ws['G1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('M1:P1')
    ws['M1'] = 'DSCR'
    ws['M1'].alignment = Alignment(horizontal='center', vertical='center')

    pst_buffer = io.BytesIO()
    wb.save(pst_buffer)
    logging.info(f"pst file saved in tmp path")
    upload_file_to_box(user_client, pst_buffer, pst_box_folder_id, pst_file_name)

    logging.info('pst comparator begin')
    pst_comparator_tmp_file = '/tmp/pst_comparator_report.xlsx'
    pst_comparator = run_look_and_clean_df(sdk, pst_comparator_look_id,'Pst Comparator')
    pst_comparator_summary = run_look_and_clean_df(sdk, pst_comparator_summary_look_id,'Pst Comparator')
    pst_comparator_summary = pst_comparator_summary.replace(np.nan, 'Grand Total')

    with pd.ExcelWriter(pst_comparator_tmp_file, engine='openpyxl') as writer:
        pst_comparator_summary.to_excel(writer, index=False, header=True, sheet_name='Summary')
        pst_comparator.to_excel(writer, index=False, header=True, sheet_name='PST Comparator')
    
    wb = load_workbook(pst_comparator_tmp_file)
    
    logging.info(f"pst comparator prep done")
    pst_comparator_buffer = io.BytesIO()
    wb.save(pst_comparator_buffer)
    upload_file_to_box(user_client, pst_comparator_buffer, pst_comp_box_folder_id, pst_comparator_file_name)

def wells_file_prep(sdk, user_client, wells_look_id, wells_comparator_look_id):

    logging.info('wells prep begin')
    wells_df = run_look_and_clean_df(sdk, wells_look_id,'Wells Ips')
    wells_df = wells_df[['Loan Count', 'Cycle', 'Fund', 'Trust', 'Series', 'Pool', 'Loan Number', 'Product Type', 'Property Type', 'FNF Loan Type', 'Expansion Scope', 'Originator', 'Property Address', 'City', 'State', 'Zip Code', 'CBSA', 'Note Date', 'First Pay Date', 'Original Maturity Date', 'Current Maturity Date', 'Extension Option Flag', 'Extension Option Amt', 'Trade Date', 'Finance Type', 'Trade Finance Buyer', 'Trade Finance Date', 'Trade Finance Waiver', 'Original Amortization Term', 'Loan Term Remaining', 'Original Principal Balance', 'UPB Actual Ending', 'Orig Max Loan', 'Max Loan', 'Trade UPB', 'Current Loan Rate', 'Property Acquire Dt', 'Property Price', 'Property Adj Price', 'LTC', 'Appraisal Amt', 'Original LTV', 'Appraisal Fixed', 'FNF Original LTV', 'Budget', 'Loan Type', 'Borrower Experience', 'Borrower Entity', 'Borrower First Name', 'Borrower Last Name', 'Co Borrower Last Name', 'Guarantors Count', 'Orig Fico Date', 'Orig Fico', 'Borrower Country', 'Credit Reserve Holdback', 'IO Strip Rate', 'Int Accrual Date', 'Trade Yield', 'Credit Reserve Holdback Percent', 'Additional Credit Reserve Holdback', 'Trade Int Accr Days', 'Trade Int Accr Amt', 'Trade Adj Cost Basis Amt', 'Trade Cap Costs', 'Trade Wire to Seller', 'UPB Actual Beginning', 'Issued Actual TDO Balance', 'Potential Gross Rent', 'Property Tax', 'Property Homeowner Ins', 'UW Net Cash', 'NCF Debt Yield', 'Rental Cash Flow Ratio', 'Occupancy Percentage', 'Lender Points', 'Cash Out Amt', 'Cross Collat', 'Recourse', 'Overall Grade', 'Credit Grade', 'Property Value Grade', 'Original Zillow Days On Market', 'Original Zillow Appreciation/(Depreciation)', 'Original Zillow Median Value', 'Current Zillow Days On Market', 'Current Zillow Appreciation/(Depreciation)', 'Current Zillow Median Value', 'Original Appraisal Date', 'Updated Appraisal or BPO Value', 'Updated Appraisal or BPO Value Date', 'Updated As-Repaired Appraisal or BPO Value', 'Issued Cut Off Date', 'Cost Basis Adjusted Ending', 'Sale Date', 'Credit Reserve Holdback Dry', 'Credit Reserve Holdback Percent Wet', 'Additional Credit Reserve Holdback Wet', 'Credit Reserve Holdback Total Wet', 'Credit Reserve Holdback Reimbursement', 'Delinquency Status', 'Days Delinquent', 'Trade Seller', 'Servicer', 'Custodian', 'Interest Accrual Method', 'Borrower Address', 'Borrower Email', 'Borrower Phone', 'Draw Escrow', 'Draw Available Count', 'Draw Available', 'Third Natural Person Borrower or Guarantor Name', 'Fourth Natural Person Borrower or Guarantor Name', 'Borrower Liquidity', 'Original Monthly Qualifying Payment', '#Months of qualifying payment in reserves', 'Subordinate Financing', 'Business Purpose Occupancy', '$Escrowed Taxes at closing', '#Months of Taxes escrowed at closing', '$Escrowed Insurance at closing', '#Months of Insurance escrowed at closing', 'T&I due within 90 days paid at closing (Y/N)', 'T&I Collected at closing sufficient to cover due dates', 'Total Monthly Gross Rent', 'Condo Eligibility', 'Monthly Tax Amount', 'Monthly Insurance Amount', 'Additional Seller Retained Interest', 'Table Funding Fee ($)', 'Stabilized Flag', 'Toorak Loan ID', 'Total Draw Amount', 'Current Maturity Date By Cycle', 'Current Loan Rate By Cycle', 'Servicer UPB']]
    logging.info(f"wells prep done")
    wells_buffer = io.BytesIO()
    wells_df.to_excel(wells_buffer, index=False, engine='openpyxl')
    upload_file_to_box(user_client, wells_buffer, wells_box_folder_id, wells_file_name)

    logging.info('wells comparator begin')
    wells_comparator = run_look_and_clean_df(sdk, wells_comparator_look_id,'Wells Comparision')
    logging.info(f"wells comparator prep done")
    wells_comparator_buffer = io.BytesIO()
    wells_comparator.to_excel(wells_comparator_buffer, index=False, engine='openpyxl')
    upload_file_to_box(user_client, wells_comparator_buffer, wells_box_folder_id, wells_comparator_file_name)
    
def opr_file_prep(sdk, user_client, opr_look_ids):


    logging.info('inside opr prep')
    tmp_file_path = '/tmp/report.xlsx'
    tmp_image_path = '/tmp/image.png'

    delinquency = run_look_and_clean_df(sdk, look_id['OPR']['delinquency'], 'Payment Status Tracker')
    del_vs_count = run_look_and_save_png(sdk, look_id['OPR']['del_vs_count'], tmp_image_path)
    maturity_date = run_look_and_clean_df(sdk, look_id['OPR']['maturity_date'], 'Payment Status Tracker')
    loan_level_opr = run_look_and_clean_df(sdk, look_id['OPR']['loan_level_opr'], 'Payment Status Tracker')

    with pd.ExcelWriter(tmp_file_path, engine = 'openpyxl') as writer:

        delinquency.to_excel(writer, sheet_name = 'Delinquency', index=False, startrow= 3)
        maturity_date.to_excel(writer, sheet_name = 'Maturity Date', index=False, startrow= 3)
        loan_level_opr.to_excel(writer, sheet_name = 'OPR - Loan Level data', index=False)


    wb = openpyxl.load_workbook(tmp_file_path)
    ws = wb['Delinquency']

    ws.merge_cells('A2:E2')
    ws['A2'] = 'Delinquency'
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws = wb['Maturity Date']
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Maturity Date'
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    wb.save(tmp_file_path)

    wb = openpyxl.load_workbook(tmp_file_path)
    wb.create_sheet("Delinquency Vs Count",2)
    ws = wb["Delinquency Vs Count"]
    img = openpyxl.drawing.image.Image(tmp_image_path)
    img.anchor = 'B2'
    img.width = 850
    ws.add_image(img)

    pst_buffer = io.BytesIO()
    wb.save(pst_buffer)
    logging.info(f"{opr_file_name} saved in tmp path")
    upload_file_to_box(user_client, pst_buffer, opr_box_folder_id, opr_file_name)

def upload_file_to_box(user_client, buffer, box_folder_id, file_name):

    logging.info(f"inside upload file function")

    try:
        
        buffer.seek(0)
        uploaded_file = user_client.folder(box_folder_id).upload_stream(buffer, file_name=file_name)
        logging.info(f"{file_name} is uploaded to box successfully")

    except Exception as e:

        logging.error(f"An error occurred: {e}")



def box_looker_conn():

    box_creds = get_secret(secret_name['box_creds'])
    box_key,box_user_id = retriveBoxConfigFromSecret(box_creds)
    config = JWTAuth.from_settings_dictionary(box_key)                    
    client = Client(config)
    user_to_impersonate = client.user(user_id=box_user_id)
    user_client = client.as_user(user_to_impersonate)

    #initiate looker sdk
    looker_creds = get_secret(secret_name['looker_creds'])
    os.environ['LOOKERSDK_BASE_URL'] = looker_creds['LOOKERSDK_BASE_URL']
    os.environ['LOOKERSDK_CLIENT_ID'] = looker_creds['LOOKERSDK_CLIENT_ID']
    os.environ['LOOKERSDK_CLIENT_SECRET'] = looker_creds['LOOKERSDK_CLIENT_SECRET']
    sdk = looker_sdk.init40()

    pst_file_prep(sdk, user_client, look_id['pst'], look_id['pst_comparator'], look_id['pst_comparator_summary'])
    time.sleep(300)
    wells_file_prep(sdk, user_client, look_id['wells_ips'], look_id['wells_comparator'])
    time.sleep(300)
    opr_file_prep(sdk, user_client, look_id['OPR'])
